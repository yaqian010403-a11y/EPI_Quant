import os
import numpy as np
import re
from PyQt5.QtWidgets import QMainWindow, QVBoxLayout, QWidget, QPushButton, QHBoxLayout, QLabel, QDialog, \
    QLineEdit, QFormLayout, QDialogButtonBox, QComboBox, QMessageBox, QSpacerItem, QSizePolicy
from PyQt5.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import pandas as pd
import time
from scipy.spatial import KDTree
from PIL import Image
from openpyxl import load_workbook







class CustomInputDialog(QDialog):
    def __init__(self, cell_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Enter Cell Information')

        self.cell_id_input = QLineEdit(self)
        self.cell_id_input.setText(cell_id)
        self.cell_id_input.setReadOnly(True)
        self.classification_input = QComboBox(self)
        self.classification_input.addItem("mes-cell", 1)
        self.classification_input.addItem("epi-cell", 0)

        layout = QFormLayout()
        layout.addRow('Cell ID:', self.cell_id_input)
        layout.addRow('Classification Results:', self.classification_input)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        layout.addWidget(button_box)
        self.setLayout(layout)

    def getInputs(self):
        return self.cell_id_input.text(), self.classification_input.currentData()


class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=10, height=8, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)

        # 鼠标事件变量
        self.drag_start = None  # 记录拖动的起始点

        # 保存初始范围
        self.initial_xlim = (-10, 10)
        self.initial_ylim = (-10, 10)
        self.ax.set_xlim(self.initial_xlim)
        self.ax.set_ylim(self.initial_ylim)

        # 绑定鼠标事件
        self.mpl_connect('button_press_event', self.on_press)
        self.mpl_connect('motion_notify_event', self.on_motion)
        self.mpl_connect('button_release_event', self.on_release)
        self.mpl_connect('scroll_event', self.on_scroll)

    def on_press(self, event):
        if event.inaxes:  # 鼠标在画布内点击
            self.drag_start = (event.xdata, event.ydata)  # 记录起始点

    def on_motion(self, event):
        if self.drag_start is None or not event.inaxes:
            return
        x_start, y_start = self.drag_start
        dx = x_start - event.xdata
        dy = y_start - event.ydata

        x_min, x_max = self.ax.get_xlim()
        y_min, y_max = self.ax.get_ylim()

        self.ax.set_xlim([x_min + dx, x_max + dx])
        self.ax.set_ylim([y_min + dy, y_max + dy])
        self.draw()

    def on_release(self, event):
        self.drag_start = None

    def on_scroll(self, event):
        x_min, x_max = self.ax.get_xlim()
        y_min, y_max = self.ax.get_ylim()

        scale_factor = 0.9 if event.button == 'up' else 1.1
        x_mid = (x_min + x_max) / 2
        y_mid = (y_min + y_max) / 2
        x_range = (x_max - x_min) * scale_factor
        y_range = (y_max - y_min) * scale_factor

        self.ax.set_xlim([x_mid - x_range / 2, x_mid + x_range / 2])
        self.ax.set_ylim([y_mid - y_range / 2, y_mid + y_range / 2])
        self.draw()

    def reset_view(self):
        self.ax.set_xlim(self.initial_xlim)
        self.ax.set_ylim(self.initial_ylim)
        self.draw()

class MainWindow(QMainWindow):
    def __init__(self, npy_folder, img_folder, output_picture, clustering_data_path, cells_info_path, output_folder, rate=0.1):
        super().__init__()
        self.setWindowTitle('Correct Cell Classification Results')
        self.setGeometry(100, 100, 800, 800)

        self.npy_folder = npy_folder
        self.img_folder = img_folder
        self.output_picture = output_picture
        self.rate = rate
        self.output_folder = output_folder
        self.cells_info = pd.read_excel(cells_info_path)
        self.cells_info_path = cells_info_path
        self.clustering_data_path = clustering_data_path
        self.clustering_data = pd.read_excel(self.clustering_data_path)


        self.clustering_data['modified'] = False
        self.cells_info['modified'] = False

        self.npy_files = sorted([f for f in os.listdir(self.npy_folder) if f.endswith('.npy')],
                                key=lambda x: int(re.findall(r'\d+', x)[0]))

        file_numbers = [int(re.findall(r'\d+', f)[0]) for f in self.npy_files]
        self.min_frame_number = min(file_numbers)
        self.max_frame_number = max(file_numbers)

        self.current_image_index = None

        self.canvas = MplCanvas(self, width=10, height=8, dpi=100)
        self.canvas.mpl_connect('button_press_event', self.on_click)

        self.canvas.ax.clear()
        self.canvas.ax.axis('off')
        self.canvas.draw()

        self.info_label = QLabel(self)
        self.info_label.setAlignment(Qt.AlignCenter)

        main_layout = QVBoxLayout()
        left_layout = QVBoxLayout()

        # 将“上一帧”和“下一帧”按钮与帧选择布局放在同一行显示
        button_and_frame_selection_layout = QHBoxLayout()

        # 添加“上一帧”和“下一帧”按钮
        prev_button = QPushButton('Previous Frame')
        prev_button.setFixedWidth(300)  # 设置按钮宽度
        prev_button.clicked.connect(self.show_previous_image)
        button_and_frame_selection_layout.addWidget(prev_button)

        # 添加“下一帧”按钮
        next_button = QPushButton('Next Frame')
        next_button.setFixedWidth(300)  # 设置按钮宽度
        next_button.clicked.connect(self.show_next_image)
        button_and_frame_selection_layout.addWidget(next_button)

        # 创建帧选择区域
        frame_selection_label = QLabel("Select Frame Number:")
        button_and_frame_selection_layout.addWidget(frame_selection_label)

        # 创建一个可编辑的下拉框来显示所有帧号，并允许用户手动输入
        self.frame_selector = QComboBox(self)
        self.frame_selector.setEditable(True)  # 允许用户手动输入帧号
        self.frame_selector.setFixedWidth(100)  # 设置下拉框宽度
        self.frame_selector.addItems([str(i) for i in range(self.min_frame_number, self.max_frame_number + 1)])
        self.frame_selector.currentTextChanged.connect(self.update_frame_input)
        button_and_frame_selection_layout.addWidget(self.frame_selector)

        # 创建“确定”按钮
        frame_select_button = QPushButton("OK")
        frame_select_button.setFixedWidth(100)  # 设置按钮宽度
        frame_select_button.clicked.connect(self.select_frame)
        button_and_frame_selection_layout.addWidget(frame_select_button)



        # 设置按钮和帧选择的间距
        button_and_frame_selection_layout.setSpacing(10)

        # 添加布局到主布局
        top_layout = QHBoxLayout()
        frame_selection_wrapper = QVBoxLayout()
        frame_selection_wrapper.addLayout(button_and_frame_selection_layout)
        frame_selection_wrapper.addStretch()
        top_layout.addLayout(frame_selection_wrapper, 1)

        left_layout.addLayout(top_layout)


        # Canvas and info label
        canvas_layout = QVBoxLayout()
        canvas_layout.setContentsMargins(0, 0, 0, 0)  # 清除canvas周围的边距
        canvas_layout.setSpacing(0)  # 控制画布和其他部分的间距
        canvas_layout.addWidget(self.canvas)
        canvas_layout.addWidget(self.info_label)
        left_layout.addLayout(canvas_layout)

        # 图例和说明区域
        legend_layout = QHBoxLayout()
        legend_layout.setContentsMargins(0, 0, 0, 0)  # 清除边距
        legend_layout.setSpacing(5)  # 移除图例与说明之间的空隙

        # 左侧空隙以使整体右对齐
        legend_layout.addStretch()

        mesophyll_color = QLabel()
        mesophyll_color.setFixedSize(60, 10)
        mesophyll_color.setStyleSheet("background-color: yellow; border: 1px solid black;")
        mesophyll_label = QLabel("mes-cell")
        legend_layout.addWidget(mesophyll_color)
        legend_layout.addWidget(mesophyll_label)

        epidermal_color = QLabel()
        epidermal_color.setFixedSize(60, 10)
        epidermal_color.setStyleSheet("background-color: blue; border: 1px solid black;")
        epidermal_label = QLabel("epi-cell")
        legend_layout.addWidget(epidermal_color)
        legend_layout.addWidget(epidermal_label)

        # 说明文字
        instruction_label = QLabel("Note: Click on the Cell ID to Modify the Cell Type")
        legend_layout.addWidget(instruction_label)

        # 添加图例布局
        left_layout.addLayout(legend_layout)

        main_layout.addLayout(left_layout)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        self.cell_contours = []

    def update_frame_input(self, text):
        """Update the frame input field based on the selected or entered frame number."""
        self.frame_selector.setCurrentText(text)

    def save_modified_rows(self, data, file_path):
        # 仅获取被修改的行
        modified_rows = data[data['modified'] == True]

        if not modified_rows.empty:
            # 打开现有的 Excel 文件
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # 根据 DataFrame 的列名获取列的顺序
            col_index = {col_name: idx + 1 for idx, col_name in enumerate(data.columns)}

            # 更新修改的行到 Excel 中
            for _, row in modified_rows.iterrows():
                excel_row = row.name + 2  # DataFrame的行索引加2，因为Excel的行从1开始，且第一行为标题
                for col_name, value in row.items():  # 将 iteritems() 改为 items()
                    if col_name == 'modified':  # 跳过标记列
                        continue
                    excel_col = col_index[col_name]
                    sheet.cell(row=excel_row, column=excel_col, value=value)

            # 保存文件
            workbook.save(file_path)
            workbook.close()

            # 更新已保存行的 'modified' 列为 False
            data.loc[modified_rows.index, 'modified'] = False

        #print(f"成功保存 {len(modified_rows)} 行修改到 {file_path}")

    def select_frame(self):
        try:
            frame_number = int(self.frame_selector.currentText())

            # 检查输入的帧号是否在范围内
            if frame_number < self.min_frame_number or frame_number > self.max_frame_number:
                QMessageBox.warning(self, 'Error',
                                    f'Frame number out of range. Please enter a number between  {self.min_frame_number} and {self.max_frame_number}.')
                return

            # 计算对应的索引
            self.current_image_index = frame_number - self.min_frame_number
            self.update_image1()  # 更新图像显示
        except ValueError:
            QMessageBox.warning(self, 'Error', 'Please enter a valid frame number.')



    def process_images(self, progress_dialog=None, show_progress_bar=None, initial_value=80):
        """绘制并保存所有图像，并动态更新进度条"""
        # 获取所有 .npy 文件的列表，并按文件名中的数字排序
        npy_files = sorted([f for f in os.listdir(self.npy_folder) if f.endswith('.npy')],
                           key=lambda x: int(re.findall(r'\d+', x)[0]))

        total_files = len(npy_files)  # 获取总文件数
        progress_step = (100 - initial_value) / total_files  # 计算每个文件的进度比例

        # 遍历每个 .npy 文件
        for i, npy_file in enumerate(npy_files):
            frame_number = int(re.findall(r'\d+', npy_file)[0])  # 从文件名提取帧号
            #print('frame_number:', frame_number)

            # 更新当前图像索引为当前的迭代索引 i
            self.current_image_index = i
            #print('current_image_index:', self.current_image_index)

            # 调用更新图像的方法
            self.update_image()

            # 显示进度信息
            show_progress_bar(f"Cell classification image for frame {frame_number} has been successfully generated...")

            # 更新进度条
            if progress_dialog:
                progress_value = initial_value + (i + 1) * progress_step
                progress_dialog.setValue(int(progress_value))

        # 打开现有的 Excel 文件
        workbook = load_workbook(self.cells_info_path)
        sheet = workbook.active

        # 确保 `cells_info` 中有 `leading_edge` 列
        if 'leading_edge' not in self.cells_info.columns:
            raise ValueError("Data does not contain the 'leading_edge' column")

        # 找到当前 Excel 文件的最后一列，并固定 leading_edge 的列位置
        last_col = sheet.max_column + 1  # 最后一列的下一个位置
        sheet.cell(row=1, column=last_col, value='leading_edge')  # 设置标题为 'leading_edge'

        # 将 `leading_edge` 列保存到 Excel 文件的最后一列
        for index, row in self.cells_info.iterrows():
            excel_row = index + 2  # Excel 行从 1 开始，第一行为标题行，所以加 2
            leading_edge_value = row['leading_edge']
            sheet.cell(row=excel_row, column=last_col, value=leading_edge_value)  # 写入最后一列

        # 保存 Excel 文件
        workbook.save(self.cells_info_path)
        workbook.close()

        # 处理完成后将进度条设置为 100
        if progress_dialog:
            progress_dialog.setValue(100)
            progress_dialog.close()

    def log_info(self, message, label):
        #print(message)
        label.setText(message)

    def show_previous_image(self):
        if self.current_image_index > 0:
            self.current_image_index -= 1
            self.update_image1()

    def show_next_image(self):
        if self.current_image_index < len(self.npy_files) - 1:
            self.current_image_index += 1
            self.update_image1()

    def jump_to_frame(self):
        try:
            frame_number = int(self.frame_input.text())  # 输入的帧号
            frame_index = next(i for i, f in enumerate(self.npy_files) if int(re.findall(r'\d+', f)[0]) == frame_number)
            self.current_image_index = frame_index
            self.update_image1()  # 更新图像显示
        except ValueError:
            QMessageBox.warning(self, 'Error', 'Please enter a valid frame number.')
        except StopIteration:
            QMessageBox.warning(self, 'Error', 'Frame number out of range. Please enter a valid frame number.')




    def get_adjacent_cells(self, blue_cells_contours, yellow_cells_contours, threshold=3):

        adjacent_pairs = []  # 存储相邻的细胞对 (大细胞ID, 小细胞ID)

        # 构建 KDTree 以便快速查找
        for large_cell_id, large_contour in yellow_cells_contours.items():
            # 使用 KDTree 来查找小细胞的轮廓点
            large_tree = KDTree(large_contour)

            # 遍历所有小细胞的轮廓，寻找相邻关系
            for small_cell_id, small_contour in blue_cells_contours.items():
                # 计算小细胞轮廓点到大细胞轮廓点的最小距离
                distances, _ = large_tree.query(small_contour, k=1)  # 查找每个小细胞轮廓点只查找它与最近的一个大细胞轮廓点之间的距离
                if np.any(distances < threshold):  # 如果存在距离小于阈值的点，则认为这两个细胞相邻
                    adjacent_pairs.append((large_cell_id, small_cell_id))  # 将相邻的细胞对加入结果

        return adjacent_pairs


    def get_matching_points_between_cells(self, large_contour, small_contour, threshold=3):

        matching_points = []  # 存储匹配的点对
        # 构建大细胞轮廓点的 KDTree
        large_tree = KDTree(large_contour)

        # 遍历小细胞的轮廓点，寻找大细胞中与之最接近的点
        for small_point in small_contour:
            distance, index = large_tree.query(small_point, k=1)
            if distance < threshold:  # 如果距离小于阈值，则记录匹配点
                matching_points.append((large_contour[index], small_point))

        return matching_points

    def update_image(self):
        #print('1', time.time())
        self.canvas.ax.clear()

        # 从文件名中提取实际的帧号
        npy_file = self.npy_files[self.current_image_index]
        #print('current_image_index', self.current_image_index)
        frame_number = int(re.findall(r'\d+', npy_file)[0])
        #print('frame_number1', frame_number)
        npy_path = os.path.join(self.npy_folder, npy_file)

        short_npy_name = os.path.basename(npy_path)
        self.log_info(f"{short_npy_name}", self.info_label)

        if os.path.exists(npy_path):
            #print('2', time.time())

            # 加载 .npy 文件
            dat = np.load(npy_path, allow_pickle=True).item()
            #print('3', time.time())

            # 提取 outlines 和 colors 数据
            outlines = dat['outlines']  # 轮廓标签图
            #base_img = dat['img']  # 使用 npy 文件中的图像
            # 如果 .npy 文件中没有 'img' 属性，加载 img_folder 中的对应图片作为 base_img
            # 检查 .npy 文件中是否包含图像
            if 'img' in dat:
                base_img = dat['img']  # 使用 npy 文件中的图像
                #print(2)

            else:
                # 如果 'img' 不存在，从 img_folder 中找到与 frame_number 完全匹配的图片
                img_file = None

                for file_name in os.listdir(self.img_folder):
                    # 提取图像文件名前的数字
                    img_frame_match = re.findall(r'\d+', file_name)
                    if img_frame_match:  # 检查是否找到了数字
                        img_frame_number = int(img_frame_match[0])  # 提取第一个连续数字
                        # 如果图像文件中的帧号与 .npy 文件中的帧号一致，并且文件是 .png 或 .img
                        if img_frame_number == frame_number and (
                                file_name.endswith('.jpeg') or file_name.endswith('.jpg') or file_name.endswith('.png')) :
                            img_file = os.path.join(self.img_folder, file_name)
                            break

                if img_file is not None:
                    # 打开图片并转换为 numpy 数组
                    base_img = np.array(Image.open(img_file))
                    #print(f"使用图片 {img_file} 作为 base_img")
                else:
                    raise FileNotFoundError(f"Image file corresponding to frame {frame_number} not found!")

            # 如果原图是灰度图，则转换为 RGB 格式
            if len(base_img.shape) == 2:
                base_img = np.stack([base_img] * 3, axis=-1)

            # 确保图像是 uint8 格式（避免数据裁剪问题）
            if base_img.dtype != np.uint8:
                base_img = (base_img / base_img.max() * 255).astype(np.uint8)

            # 创建一个 RGB 图像的副本
            img = base_img.copy()

            # 构建一个 Cell_Index 到 Cluster_Label 的映射字典
            cell_type_dict = self.clustering_data.set_index('Cell_Index')['Cluster_Label'].to_dict()

            # 存储细胞轮廓，用于邻接性分析
            blue_cells_contours = {}  # 小细胞（蓝色细胞）
            yellow_cells_contours = {}  # 大细胞（黄色细胞）

            # 遍历每个不同的轮廓 ID，并根据 Excel 中的类别进行填充
            unique_ids = np.unique(outlines)  # 提取唯一的轮廓 ID
            for cell_id in unique_ids:
                if cell_id == 0:
                    continue  # 忽略背景

                # 查找当前 cell_id 的类别
                cell_index = f'Cell{frame_number}_{cell_id}'  # 按照 '细胞{frame_number}_{cell_id}' 格式构建索引
                cluster_label = cell_type_dict.get(cell_index, -1)  # 如果未找到类别，返回 -1

                # 根据 Cluster_Label 选择填充颜色，并记录细胞轮廓
                if cluster_label == 1:
                    color = (255, 255, 0)  # 黄色（大细胞）
                    yellow_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # 存储大细胞轮廓
                elif cluster_label == 0:
                    color = (0, 0, 255)  # 蓝色（小细胞）
                    blue_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # 存储小细胞轮廓
                else:
                    continue  # 跳过其他类别

                # 将该轮廓 ID 的所有像素位置标记为指定颜色
                img[outlines == cell_id] = color

            #print('5', time.time())

            # 找到相邻的大细胞和小细胞的细胞对
            adjacent_pairs = self.get_adjacent_cells(blue_cells_contours, yellow_cells_contours, threshold=3)
            #print(f"Adjacent cell pairs: {adjacent_pairs}")




            # 在 Matplotlib 中显示图像
            ax = self.canvas.ax
            ax.imshow(img)
            ax.axis('off')  # 移除坐标轴
            self.canvas.fig.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            ax.set_position([0, 0, 1, 1])

            # 所有相邻大细胞和小细胞之间的分隔点集合
            all_matching_points = []

            # 标记每个相邻细胞对并绘制分隔线
            for large_cell_id, small_cell_id in adjacent_pairs:
                # 找到相邻的匹配点
                large_contour = yellow_cells_contours[large_cell_id]
                small_contour = blue_cells_contours[small_cell_id]

                # 获取每对相邻细胞之间的所有匹配点
                matching_points = self.get_matching_points_between_cells(large_contour, small_contour, threshold=3)

                # 将所有匹配点加入总的点集中
                all_matching_points.extend([pt for large_point, pt in matching_points])

            # 使用 KDTree 来计算每个当前帧细胞到所有匹配点的最短距离
            if all_matching_points:
                tree = KDTree(all_matching_points)

                # 遍历 cells_info 中的所有细胞，过滤出当前帧的细胞
                #print(f"Adjacent pairs for frame {frame_number}: {adjacent_pairs}")

                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # 获取细胞编号
                    cluster_label = cell_type_dict.get(cell_index, -1)  # 从 cell_type_dict 中查找类别

                    # 如果该细胞不属于当前帧，跳过
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue
                    #print('细胞编号', cell_index)


                    # 如果类别不在 [0, 1] 中，跳过
                    if cluster_label not in [0, 1]:
                        continue

                    # 计算该细胞的中心点
                    cell_positions = np.where(outlines == int(cell_index.split('_')[-1]))
                    center_y = int(np.mean(cell_positions[0]))  # 计算 y 坐标的平均值
                    center_x = int(np.mean(cell_positions[1]))  # 计算 x 坐标的平均值
                    cell_center = np.array([center_y, center_x])

                    # 计算到所有分隔点的最短距离
                    distance, _ = tree.query(cell_center)

                    self.cells_info.at[index, 'leading_edge'] = distance

                    # 根据细胞类型决定 leading_edge 符号
                    #if cluster_label == 1:  # 大细胞
                    #    self.cells_info.at[index, 'leading_edge'] = -distance
                    #else:  # 小细胞
                    #    self.cells_info.at[index, 'leading_edge'] = distance

                    #if cluster_label == 1:  # 大细胞
                    #    self.update_cells_info(cell_index, -distance)
                    #else:  # 小细胞
                    #    self.update_cells_info(cell_index, distance)



            else:
                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # 获取细胞编号
                    cluster_label = cell_type_dict.get(cell_index, -1)  # 从 cell_type_dict 中查找类别

                    # 如果该细胞不属于当前帧，跳过
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue
                    #print('细胞编号', cell_index)


                    # 如果类别不在 [0, 1] 中，跳过
                    if cluster_label not in [0, 1]:
                        continue

                    # 根据细胞类型决定 leading_edge 符号
                    if cluster_label == 1:  # 大细胞
                        self.cells_info.at[index, 'leading_edge'] = -1
                    else:  # 小细胞
                        self.cells_info.at[index, 'leading_edge'] = 1

                    # 计算 leading_edge 并更新 cells_info 的值
                    #if cluster_label == 1:  # 大细胞
                    #    self.update_cells_info(cell_index, -1)
                    #else:  # 小细胞
                    #    self.update_cells_info(cell_index, 1)


            # 绘制完成后刷新图像
            #print('6', time.time())
            self.canvas.draw()
            #print('6', time.time())


            # Pillow 保存图片
            output_dir = self.output_picture
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            save_path = os.path.join(output_dir, f'{frame_number}_leading_edge.png')
            pil_img = Image.fromarray(img)
            pil_img.save(save_path)  # 使用 Pillow 保存图像
            #print('9', time.time())
            #print(f"Image saved successfully at {save_path}")

    def update_image1(self):
        #print('1', time.time())
        self.canvas.ax.clear()

        # 从文件名中提取实际的帧号
        npy_file = self.npy_files[self.current_image_index]
        frame_number = int(re.findall(r'\d+', npy_file)[0])
        #print(frame_number)
        npy_path = os.path.join(self.npy_folder, npy_file)

        short_npy_name = os.path.basename(npy_path)
        self.log_info(f"Loading: {short_npy_name}", self.info_label)

        if os.path.exists(npy_path):
            #print('2', time.time())

            # 加载 .npy 文件
            dat = np.load(npy_path, allow_pickle=True).item()
            #print('3', time.time())

            # 提取 outlines 和 colors 数据
            outlines = dat['outlines']  # 轮廓标签图
            #base_img = dat['img']  # 使用 npy 文件中的图像

            if 'img' in dat:
                base_img = dat['img']  # 使用 npy 文件中的图像
            else:
                # 如果 'img' 不存在，从 img_folder 中找到与 frame_number 完全匹配的图片
                img_file = None

                for file_name in os.listdir(self.img_folder):
                    # 提取图像文件名前的数字
                    img_frame_match = re.findall(r'\d+', file_name)
                    if img_frame_match:  # 检查是否找到了数字
                        img_frame_number = int(img_frame_match[0])  # 提取第一个连续数字
                        # 如果图像文件中的帧号与 .npy 文件中的帧号一致，并且文件是 .png 或 .img
                        if img_frame_number == frame_number and (
                                file_name.endswith('.jpeg') or file_name.endswith('.jpg') or file_name.endswith('.png')):
                            img_file = os.path.join(self.img_folder, file_name)
                            break

                if img_file is not None:
                    # 打开图片并转换为 numpy 数组
                    base_img = np.array(Image.open(img_file))
                    #print(f"使用图片 {img_file} 作为 base_img")
                else:
                    raise FileNotFoundError(f"Image file corresponding to frame {frame_number} not found!")


            # 如果原图是灰度图，则转换为 RGB 格式
            if len(base_img.shape) == 2:
                base_img = np.stack([base_img] * 3, axis=-1)

            # 确保图像是 uint8 格式（避免数据裁剪问题）
            if base_img.dtype != np.uint8:
                base_img = (base_img / base_img.max() * 255).astype(np.uint8)

            # 创建一个 RGB 图像的副本
            img = base_img.copy()

            # 构建一个 Cell_Index 到 Cluster_Label 的映射字典
            cell_type_dict = self.clustering_data.set_index('Cell_Index')['Cluster_Label'].to_dict()

            # 存储细胞轮廓，用于邻接性分析
            blue_cells_contours = {}  # 小细胞（蓝色细胞）
            yellow_cells_contours = {}  # 大细胞（黄色细胞）

            # 遍历每个不同的轮廓 ID，并根据 Excel 中的类别进行填充
            unique_ids = np.unique(outlines)  # 提取唯一的轮廓 ID
            for cell_id in unique_ids:
                if cell_id == 0:
                    continue  # 忽略背景

                # 查找当前 cell_id 的类别
                cell_index = f'Cell{frame_number}_{cell_id}'  # 按照 '细胞{frame_number}_{cell_id}' 格式构建索引
                cluster_label = cell_type_dict.get(cell_index, -1)  # 如果未找到类别，返回 -1

                # 根据 Cluster_Label 选择填充颜色，并记录细胞轮廓
                if cluster_label == 1:
                    color = (255, 255, 0)  # 黄色（大细胞）
                    yellow_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # 存储大细胞轮廓
                elif cluster_label == 0:
                    color = (0, 0, 255)  # 蓝色（小细胞）
                    blue_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # 存储小细胞轮廓
                else:
                    continue  # 跳过其他类别

                # 将该轮廓 ID 的所有像素位置标记为指定颜色
                img[outlines == cell_id] = color

            #print('5', time.time())

            # 找到相邻的大细胞和小细胞的细胞对
            adjacent_pairs = self.get_adjacent_cells(blue_cells_contours, yellow_cells_contours, threshold=3)
            #print(f"Adjacent cell pairs: {adjacent_pairs}")

            # 在 Matplotlib 中显示图像
            ax = self.canvas.ax
            ax.imshow(img)
            ax.axis('off')  # 移除坐标轴
            self.canvas.fig.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            ax.set_position([0, 0, 1, 1])

            # 所有相邻大细胞和小细胞之间的分隔点集合
            all_matching_points = []

            #print('6', time.time())

            # 标记每个相邻细胞对并绘制分隔线
            for large_cell_id, small_cell_id in adjacent_pairs:
                # 找到相邻的匹配点
                large_contour = yellow_cells_contours[large_cell_id]
                small_contour = blue_cells_contours[small_cell_id]

                # 获取每对相邻细胞之间的所有匹配点
                matching_points = self.get_matching_points_between_cells(large_contour, small_contour, threshold=3)

                # 将所有匹配点加入总的点集中
                all_matching_points.extend([pt for large_point, pt in matching_points])

                # 绘制相邻点之间的分隔线
                for (large_point, small_point) in matching_points:
                    ax.plot([large_point[1], small_point[1]], [large_point[0], small_point[0]], color='red',
                            linestyle='--')
            #print('7', time.time())

            # 添加细胞 ID 标注
            for i, cell_id in enumerate(unique_ids):
                if cell_id == 0:
                    continue  # 忽略背景

                # 查找当前 cell_id 的类别，仅标注大细胞和小细胞的 ID
                cell_index = f'Cell{frame_number}_{cell_id}'
                cluster_label = cell_type_dict.get(cell_index, -1)

                # 仅标注大细胞和小细胞的 ID，跳过其他类别
                if cluster_label not in [0, 1]:
                    continue

                # 获取当前细胞 ID 的所有像素位置 (y, x)
                positions = np.where(outlines == cell_id)

                # 计算细胞的中心位置
                center_y = int(np.mean(positions[0]))  # 计算 y 坐标的平均值
                center_x = int(np.mean(positions[1]))  # 计算 x 坐标的平均值

                # 根据细胞类别设置文本框的颜色
                if cluster_label == 1:  # 大细胞（黄色标记）
                    bbox_props = dict(facecolor='yellow', alpha=0.5, edgecolor='none')
                elif cluster_label == 0:  # 小细胞（蓝色标记）
                    bbox_props = dict(facecolor='blue', alpha=0.5, edgecolor='none')
                else:
                    continue

                # 标注细胞 ID 到对应的中心位置
                ax.text(center_x, center_y, str(cell_id), color='white', fontsize=7, bbox=bbox_props)

            #print('8', time.time())

            # 使用 KDTree 来计算每个当前帧细胞到所有匹配点的最短距离
            if all_matching_points:
                tree = KDTree(all_matching_points)

                # 遍历 cells_info 中的所有细胞，过滤出当前帧的细胞

                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # 获取细胞编号
                    cluster_label = cell_type_dict.get(cell_index, -1)  # 从 cell_type_dict 中查找类别

                    # 如果该细胞不属于当前帧，跳过
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue


                    # 如果类别不在 [0, 1] 中，跳过
                    if cluster_label not in [0, 1]:
                        continue

                    # 计算该细胞的中心点
                    cell_positions = np.where(outlines == int(cell_index.split('_')[-1]))
                    center_y = int(np.mean(cell_positions[0]))  # 计算 y 坐标的平均值
                    center_x = int(np.mean(cell_positions[1]))  # 计算 x 坐标的平均值
                    cell_center = np.array([center_y, center_x])

                    # 计算到所有分隔点的最短距离
                    distance, _ = tree.query(cell_center)

                    # 根据细胞类型决定 leading_edge 符号
                    #if cluster_label == 1:  # 大细胞
                    #    self.cells_info.at[index, 'leading_edge'] = -distance
                    #else:  # 小细胞
                    #    self.cells_info.at[index, 'leading_edge'] = distance

                    #if cluster_label == 1:  # 大细胞
                    #    #print(cell_index)
                    #    self.update_cells_info(cell_index, -distance)
                    #else:  # 小细胞
                    #    #print(cell_index)
                    #    self.update_cells_info(cell_index, distance)

                    self.update_cells_info(cell_index, distance)

            # 绘制完成后刷新图像
            self.canvas.draw()
            #print('9', time.time())

            # 保存 Excel

            #print('9', time.time())

            # Pillow 保存图片
            output_dir = self.output_picture
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            # 保存 PDF 图像（追加）
            #pdf_save_path = os.path.join(output_dir, f'{frame_number}_leading_edge.pdf')
            #self.canvas.fig.savefig(pdf_save_path, format='pdf', bbox_inches='tight', pad_inches=0, dpi=300)

            save_path = os.path.join(output_dir, f'{frame_number}_leading_edge.png')
            #self.cells_info.to_excel(self.cells_info_path, index=False)
            pil_img = Image.fromarray(img)
            pil_img.save(save_path)  # 使用 Pillow 保存图像
            #print('10', time.time())
            #print(f"Image saved successfully at {save_path} using Pillow!")



    def on_click(self, event):
        if event.inaxes:
            for text in event.inaxes.texts:
                contains, _ = text.contains(event)
                if contains:
                    cell_id = text.get_text()
                    cell_index = f'Cell{self.current_image_index + 1}_{cell_id}'
                    dialog = CustomInputDialog(cell_index, self)
                    if dialog.exec_() == QDialog.Accepted:
                        new_cell_id, classification_result = dialog.getInputs()
                        #print(f"Cell ID updated to: {new_cell_id}, Classification result: {classification_result}")
                        # 更新聚类数据并保存
                        self.update_clustering_data(cell_index, classification_result)
                        self.save_clustering_data()

                    break

    def update_clustering_data(self, cell_index, classification_result):
        row = self.clustering_data[self.clustering_data['Cell_Index'] == cell_index]
        if not row.empty:
            #print(
                #f"更新前: {cell_index} 的分类结果为: {self.clustering_data.loc[self.clustering_data['Cell_Index'] == cell_index, 'Cluster_Label'].values[0]}")

            # 更新分类结果并标记该行已修改
            if classification_result == 1:  # 更新为大细胞
                self.clustering_data.loc[
                    self.clustering_data['Cell_Index'] == cell_index, ['Cluster_Label', 'Cell_Type']] = [1, 'mes Cell']
            elif classification_result == 0:  # 更新为小细胞
                self.clustering_data.loc[
                    self.clustering_data['Cell_Index'] == cell_index, ['Cluster_Label', 'Cell_Type']] = [0, 'epi Cell']

            # 标记为已修改
            self.clustering_data.loc[self.clustering_data['Cell_Index'] == cell_index, 'modified'] = True

    def save_clustering_data(self):
        try:
            self.update_image1()
            # 保存被修改的行
            self.save_modified_rows(self.clustering_data, self.clustering_data_path)



            self.save_modified_rows(self.cells_info, self.cells_info_path)
            QMessageBox.information(self, 'Save Successful', 'Changes have been saved to the file.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred while saving: {str(e)}')

    def update_cells_info(self, cell_index, new_value):
        # 检查 'leading_edge' 列是否存在，如果不存在则添加
        if 'leading_edge' not in self.cells_info.columns:
            self.cells_info['leading_edge'] = None  # 添加新列，默认值为 None 或你想要的其他默认值

        # 查找匹配的细胞编号
        row = self.cells_info[self.cells_info['Cell Index'] == cell_index]
        if not row.empty:
            # 更新 'leading_edge' 列的值
            self.cells_info.loc[self.cells_info['Cell Index'] == cell_index, 'leading_edge'] = new_value

            # 如果 'modified' 列不存在，先添加
            if 'modified' not in self.cells_info.columns:
                self.cells_info['modified'] = False  # 添加新列，默认值为 False

            # 标记为已修改
            self.cells_info.loc[self.cells_info['Cell Index'] == cell_index, 'modified'] = True



