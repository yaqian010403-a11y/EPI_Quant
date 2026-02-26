import os
import numpy as np
import re
from PyQt5.QtWidgets import QMainWindow, QVBoxLayout, QWidget, QPushButton, QHBoxLayout, QLabel, QDialog, \
    QLineEdit, QFormLayout, QDialogButtonBox, QComboBox, QMessageBox, QSpacerItem, QSizePolicy
from PyQt5.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import pandas as pd
import time
from scipy.spatial import KDTree
from PIL import Image
from openpyxl import load_workbook


class CustomInputDialog(QDialog):
    def __init__(self, cell_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Enter Cell Information')

        self.cell_id_input = QLineEdit(self)
        self.cell_id_input.setText(cell_id)
        self.cell_id_input.setReadOnly(True)
        self.classification_input = QComboBox(self)
        self.classification_input.addItem("mes-cell", 1)
        self.classification_input.addItem("epi-cell", 0)

        layout = QFormLayout()
        layout.addRow('Cell ID:', self.cell_id_input)
        layout.addRow('Classification Results:', self.classification_input)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)

        layout.addWidget(button_box)
        self.setLayout(layout)

    def getInputs(self):
        return self.cell_id_input.text(), self.classification_input.currentData()


class MplCanvas(FigureCanvas):
    def __init__(self, parent=None, width=10, height=8, dpi=100):
        self.fig = Figure(figsize=(width, height), dpi=dpi)
        self.ax = self.fig.add_subplot(111)
        super().__init__(self.fig)

        # Mouse event variables
        self.drag_start = None  # Record the starting point of dragging

        # Save initial range
        self.initial_xlim = (-10, 10)
        self.initial_ylim = (-10, 10)
        self.ax.set_xlim(self.initial_xlim)
        self.ax.set_ylim(self.initial_ylim)

        # Bind mouse events
        self.mpl_connect('button_press_event', self.on_press)
        self.mpl_connect('motion_notify_event', self.on_motion)
        self.mpl_connect('button_release_event', self.on_release)
        self.mpl_connect('scroll_event', self.on_scroll)

    def on_press(self, event):
        if event.inaxes:  # Mouse click within canvas
            self.drag_start = (event.xdata, event.ydata)  # Record starting point

    def on_motion(self, event):
        if self.drag_start is None or not event.inaxes:
            return
        x_start, y_start = self.drag_start
        dx = x_start - event.xdata
        dy = y_start - event.ydata

        x_min, x_max = self.ax.get_xlim()
        y_min, y_max = self.ax.get_ylim()

        self.ax.set_xlim([x_min + dx, x_max + dx])
        self.ax.set_ylim([y_min + dy, y_max + dy])
        self.draw()

    def on_release(self, event):
        self.drag_start = None

    def on_scroll(self, event):
        x_min, x_max = self.ax.get_xlim()
        y_min, y_max = self.ax.get_ylim()

        scale_factor = 0.9 if event.button == 'up' else 1.1
        x_mid = (x_min + x_max) / 2
        y_mid = (y_min + y_max) / 2
        x_range = (x_max - x_min) * scale_factor
        y_range = (y_max - y_min) * scale_factor

        self.ax.set_xlim([x_mid - x_range / 2, x_mid + x_range / 2])
        self.ax.set_ylim([y_mid - y_range / 2, y_mid + y_range / 2])
        self.draw()

    def reset_view(self):
        self.ax.set_xlim(self.initial_xlim)
        self.ax.set_ylim(self.initial_ylim)
        self.draw()

class MainWindow(QMainWindow):
    def __init__(self, npy_folder, img_folder, output_picture, clustering_data_path, cells_info_path, output_folder, rate=0.1):
        super().__init__()
        self.setWindowTitle('Correct Cell Classification Results')
        self.setGeometry(100, 100, 800, 800)

        self.npy_folder = npy_folder
        self.img_folder = img_folder
        self.output_picture = output_picture
        self.rate = rate
        self.output_folder = output_folder
        self.cells_info = pd.read_excel(cells_info_path)
        self.cells_info_path = cells_info_path
        self.clustering_data_path = clustering_data_path
        self.clustering_data = pd.read_excel(self.clustering_data_path)

        self.clustering_data['modified'] = False
        self.cells_info['modified'] = False

        self.npy_files = sorted([f for f in os.listdir(self.npy_folder) if f.endswith('.npy')],
                                key=lambda x: int(re.findall(r'\d+', x)[0]))

        file_numbers = [int(re.findall(r'\d+', f)[0]) for f in self.npy_files]
        self.min_frame_number = min(file_numbers)
        self.max_frame_number = max(file_numbers)

        self.current_image_index = None

        self.canvas = MplCanvas(self, width=10, height=8, dpi=100)
        self.canvas.mpl_connect('button_press_event', self.on_click)

        self.canvas.ax.clear()
        self.canvas.ax.axis('off')
        self.canvas.draw()

        self.info_label = QLabel(self)
        self.info_label.setAlignment(Qt.AlignCenter)

        main_layout = QVBoxLayout()
        left_layout = QVBoxLayout()

        # Place "Previous Frame" and "Next Frame" buttons and frame selection layout on the same row
        button_and_frame_selection_layout = QHBoxLayout()

        # Add "Previous Frame" button
        prev_button = QPushButton('Previous Frame')
        prev_button.setFixedWidth(300)  # Set button width
        prev_button.clicked.connect(self.show_previous_image)
        button_and_frame_selection_layout.addWidget(prev_button)

        # Add "Next Frame" button
        next_button = QPushButton('Next Frame')
        next_button.setFixedWidth(300)  # Set button width
        next_button.clicked.connect(self.show_next_image)
        button_and_frame_selection_layout.addWidget(next_button)

        # Create frame selection area
        frame_selection_label = QLabel("Select Frame Number:")
        button_and_frame_selection_layout.addWidget(frame_selection_label)

        # Create an editable combo box to display all frame numbers, allowing manual input
        self.frame_selector = QComboBox(self)
        self.frame_selector.setEditable(True)  # Allow manual input
        self.frame_selector.setFixedWidth(100)  # Set combo box width
        self.frame_selector.addItems([str(i) for i in range(self.min_frame_number, self.max_frame_number + 1)])
        self.frame_selector.currentTextChanged.connect(self.update_frame_input)
        button_and_frame_selection_layout.addWidget(self.frame_selector)

        # Create "OK" button
        frame_select_button = QPushButton("OK")
        frame_select_button.setFixedWidth(100)  # Set button width
        frame_select_button.clicked.connect(self.select_frame)
        button_and_frame_selection_layout.addWidget(frame_select_button)

        # Set spacing for buttons and frame selection
        button_and_frame_selection_layout.setSpacing(10)

        # Add layout to main layout
        top_layout = QHBoxLayout()
        frame_selection_wrapper = QVBoxLayout()
        frame_selection_wrapper.addLayout(button_and_frame_selection_layout)
        frame_selection_wrapper.addStretch()
        top_layout.addLayout(frame_selection_wrapper, 1)

        left_layout.addLayout(top_layout)

        # Canvas and info label
        canvas_layout = QVBoxLayout()
        canvas_layout.setContentsMargins(0, 0, 0, 0)  # Clear margins around canvas
        canvas_layout.setSpacing(0)  # Control spacing between canvas and other elements
        canvas_layout.addWidget(self.canvas)
        canvas_layout.addWidget(self.info_label)
        left_layout.addLayout(canvas_layout)

        # Legend and instruction area
        legend_layout = QHBoxLayout()
        legend_layout.setContentsMargins(0, 0, 0, 0)  # Clear margins
        legend_layout.setSpacing(5)  # Remove gap between legend and instructions

        # Left spacing to align right
        legend_layout.addStretch()

        mesophyll_color = QLabel()
        mesophyll_color.setFixedSize(60, 10)
        mesophyll_color.setStyleSheet("background-color: yellow; border: 1px solid black;")
        mesophyll_label = QLabel("mes-cell")
        legend_layout.addWidget(mesophyll_color)
        legend_layout.addWidget(mesophyll_label)

        epidermal_color = QLabel()
        epidermal_color.setFixedSize(60, 10)
        epidermal_color.setStyleSheet("background-color: blue; border: 1px solid black;")
        epidermal_label = QLabel("epi-cell")
        legend_layout.addWidget(epidermal_color)
        legend_layout.addWidget(epidermal_label)

        # Instruction text
        instruction_label = QLabel("Note: Click on the Cell ID to Modify the Cell Type")
        legend_layout.addWidget(instruction_label)

        # Add legend layout
        left_layout.addLayout(legend_layout)

        main_layout.addLayout(left_layout)

        container = QWidget()
        container.setLayout(main_layout)
        self.setCentralWidget(container)

        self.cell_contours = []

    def update_frame_input(self, text):
        """Update the frame input field based on the selected or entered frame number."""
        self.frame_selector.setCurrentText(text)

    def save_modified_rows(self, data, file_path):
        # Get only modified rows
        modified_rows = data[data['modified'] == True]

        if not modified_rows.empty:
            # Open existing Excel file
            workbook = load_workbook(file_path)
            sheet = workbook.active

            # Get column order based on DataFrame column names
            col_index = {col_name: idx + 1 for idx, col_name in enumerate(data.columns)}

            # Update modified rows to Excel
            for _, row in modified_rows.iterrows():
                excel_row = row.name + 2  # DataFrame row index plus 2, Excel starts at 1 and first row is header
                for col_name, value in row.items():
                    if col_name == 'modified':  # Skip the modified flag column
                        continue
                    excel_col = col_index[col_name]
                    sheet.cell(row=excel_row, column=excel_col, value=value)

            # Save file
            workbook.save(file_path)
            workbook.close()

            # Update 'modified' column to False for saved rows
            data.loc[modified_rows.index, 'modified'] = False

    def select_frame(self):
        try:
            frame_number = int(self.frame_selector.currentText())

            # Check if input frame number is within range
            if frame_number < self.min_frame_number or frame_number > self.max_frame_number:
                QMessageBox.warning(self, 'Error',
                                    f'Frame number out of range. Please enter a number between {self.min_frame_number} and {self.max_frame_number}.')
                return

            # Calculate corresponding index
            self.current_image_index = frame_number - self.min_frame_number
            self.update_image1()  # Update image display
        except ValueError:
            QMessageBox.warning(self, 'Error', 'Please enter a valid frame number.')

    def process_images(self, progress_dialog=None, show_progress_bar=None, initial_value=80):
        """Draw and save all images, dynamically updating progress bar"""
        # Get list of all .npy files sorted by numbers in filenames
        npy_files = sorted([f for f in os.listdir(self.npy_folder) if f.endswith('.npy')],
                           key=lambda x: int(re.findall(r'\d+', x)[0]))

        total_files = len(npy_files)  # Get total number of files
        progress_step = (100 - initial_value) / total_files  # Calculate progress increment per file

        # Iterate through each .npy file
        for i, npy_file in enumerate(npy_files):
            frame_number = int(re.findall(r'\d+', npy_file)[0])  # Extract frame number from filename

            # Update current image index to current iteration index i
            self.current_image_index = i

            # Call update image method
            self.update_image()

            # Display progress message
            show_progress_bar(f"Cell classification image for frame {frame_number} has been successfully generated...")

            # Update progress bar
            if progress_dialog:
                progress_value = initial_value + (i + 1) * progress_step
                progress_dialog.setValue(int(progress_value))

        # Open existing Excel file
        workbook = load_workbook(self.cells_info_path)
        sheet = workbook.active

        # Ensure cells_info has 'leading_edge' column
        if 'leading_edge' not in self.cells_info.columns:
            raise ValueError("Data does not contain the 'leading_edge' column")

        # Find the last column in current Excel file and fix leading_edge column position
        last_col = sheet.max_column + 1  # Next position after last column
        sheet.cell(row=1, column=last_col, value='leading_edge')  # Set header to 'leading_edge'

        # Save leading_edge column to last column of Excel file
        for index, row in self.cells_info.iterrows():
            excel_row = index + 2  # Excel rows start at 1, first row is header, so add 2
            leading_edge_value = row['leading_edge']
            sheet.cell(row=excel_row, column=last_col, value=leading_edge_value)  # Write to last column

        # Save Excel file
        workbook.save(self.cells_info_path)
        workbook.close()

        # Set progress bar to 100 after processing
        if progress_dialog:
            progress_dialog.setValue(100)
            progress_dialog.close()

    def log_info(self, message, label):
        label.setText(message)

    def show_previous_image(self):
        if self.current_image_index > 0:
            self.current_image_index -= 1
            self.update_image1()

    def show_next_image(self):
        if self.current_image_index < len(self.npy_files) - 1:
            self.current_image_index += 1
            self.update_image1()

    def jump_to_frame(self):
        try:
            frame_number = int(self.frame_input.text())  # Input frame number
            frame_index = next(i for i, f in enumerate(self.npy_files) if int(re.findall(r'\d+', f)[0]) == frame_number)
            self.current_image_index = frame_index
            self.update_image1()  # Update image display
        except ValueError:
            QMessageBox.warning(self, 'Error', 'Please enter a valid frame number.')
        except StopIteration:
            QMessageBox.warning(self, 'Error', 'Frame number out of range. Please enter a valid frame number.')

    def get_adjacent_cells(self, blue_cells_contours, yellow_cells_contours, threshold=3):
        adjacent_pairs = []  # Store adjacent cell pairs (large_cell_id, small_cell_id)

        # Build KDTree for fast lookup
        for large_cell_id, large_contour in yellow_cells_contours.items():
            # Use KDTree to find small cell contour points
            large_tree = KDTree(large_contour)

            # Iterate through all small cell contours to find adjacency
            for small_cell_id, small_contour in blue_cells_contours.items():
                # Calculate minimum distance from small cell contour points to large cell contour points
                distances, _ = large_tree.query(small_contour, k=1)  # Find minimum distance
                if np.any(distances < threshold):  # If distance less than threshold, cells are adjacent
                    adjacent_pairs.append((large_cell_id, small_cell_id))  # Add adjacent cell pair

        return adjacent_pairs

    def get_matching_points_between_cells(self, large_contour, small_contour, threshold=3):
        matching_points = []  # Store matching point pairs
        # Build KDTree of large cell contour points
        large_tree = KDTree(large_contour)

        # Iterate through small cell contour points to find closest points in large cell
        for small_point in small_contour:
            distance, index = large_tree.query(small_point, k=1)
            if distance < threshold:  # If distance is below threshold, record matching points
                matching_points.append((large_contour[index], small_point))

        return matching_points

    def update_image(self):
        self.canvas.ax.clear()

        # Extract actual frame number from filename
        npy_file = self.npy_files[self.current_image_index]
        frame_number = int(re.findall(r'\d+', npy_file)[0])
        npy_path = os.path.join(self.npy_folder, npy_file)

        short_npy_name = os.path.basename(npy_path)
        self.log_info(f"{short_npy_name}", self.info_label)

        if os.path.exists(npy_path):
            # Load .npy file
            dat = np.load(npy_path, allow_pickle=True).item()

            # Extract outlines and colors data
            outlines = dat['outlines']  # Contour label map

            # Check if image is in .npy file
            if 'img' in dat:
                base_img = dat['img']  # Use image from npy file
            else:
                # If 'img' not present, find corresponding image from img_folder matching frame_number
                img_file = None

                for file_name in os.listdir(self.img_folder):
                    # Extract numbers from image filename
                    img_frame_match = re.findall(r'\d+', file_name)
                    if img_frame_match:  # Check if numbers found
                        img_frame_number = int(img_frame_match[0])  # Extract first continuous number
                        # If image file frame matches npy file frame and is .png or .jpg
                        if img_frame_number == frame_number and (
                                file_name.endswith('.jpeg') or file_name.endswith('.jpg') or file_name.endswith('.png')):
                            img_file = os.path.join(self.img_folder, file_name)
                            break

                if img_file is not None:
                    # Open image and convert to numpy array
                    base_img = np.array(Image.open(img_file))
                else:
                    raise FileNotFoundError(f"Image file corresponding to frame {frame_number} not found!")

            # Convert grayscale image to RGB format
            if len(base_img.shape) == 2:
                base_img = np.stack([base_img] * 3, axis=-1)

            # Ensure image is uint8 format (avoid data clipping issues)
            if base_img.dtype != np.uint8:
                base_img = (base_img / base_img.max() * 255).astype(np.uint8)

            # Create a copy of RGB image
            img = base_img.copy()

            # Build mapping dictionary from Cell_Index to Cluster_Label
            cell_type_dict = self.clustering_data.set_index('Cell_Index')['Cluster_Label'].to_dict()

            # Store cell contours for adjacency analysis
            blue_cells_contours = {}  # Small cells (blue cells)
            yellow_cells_contours = {}  # Large cells (yellow cells)

            # Iterate through each unique contour ID and fill based on Excel category
            unique_ids = np.unique(outlines)  # Extract unique contour IDs
            for cell_id in unique_ids:
                if cell_id == 0:
                    continue  # Ignore background

                # Find category for current cell_id
                cell_index = f'Cell{frame_number}_{cell_id}'  # Build index in format Cell{frame_number}_{cell_id}
                cluster_label = cell_type_dict.get(cell_index, -1)  # Return -1 if category not found

                # Select fill color based on Cluster_Label and record cell contours
                if cluster_label == 1:
                    color = (255, 255, 0)  # Yellow (large cell)
                    yellow_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # Store large cell contour
                elif cluster_label == 0:
                    color = (0, 0, 255)  # Blue (small cell)
                    blue_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # Store small cell contour
                else:
                    continue  # Skip other categories

                # Mark all pixel positions of this contour ID with specified color
                img[outlines == cell_id] = color

            # Find adjacent large and small cell pairs
            adjacent_pairs = self.get_adjacent_cells(blue_cells_contours, yellow_cells_contours, threshold=3)

            # Display image in Matplotlib
            ax = self.canvas.ax
            ax.imshow(img)
            ax.axis('off')  # Remove axes
            self.canvas.fig.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            ax.set_position([0, 0, 1, 1])

            # Collection of all separation points between adjacent large and small cells
            all_matching_points = []

            # Mark each adjacent cell pair and draw separation lines
            for large_cell_id, small_cell_id in adjacent_pairs:
                # Find adjacent matching points
                large_contour = yellow_cells_contours[large_cell_id]
                small_contour = blue_cells_contours[small_cell_id]

                # Get all matching points between this pair of adjacent cells
                matching_points = self.get_matching_points_between_cells(large_contour, small_contour, threshold=3)

                # Add all matching points to total point collection
                all_matching_points.extend([pt for large_point, pt in matching_points])

            # Use KDTree to calculate shortest distance from each cell in current frame to all matching points
            if all_matching_points:
                tree = KDTree(all_matching_points)

                # Iterate through all cells in cells_info, filtering for current frame cells
                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # Get cell number
                    cluster_label = cell_type_dict.get(cell_index, -1)  # Look up category from cell_type_dict

                    # If cell doesn't belong to current frame, skip
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue

                    # If category not in [0, 1], skip
                    if cluster_label not in [0, 1]:
                        continue

                    # Calculate cell center point
                    cell_positions = np.where(outlines == int(cell_index.split('_')[-1]))
                    center_y = int(np.mean(cell_positions[0]))  # Calculate average y coordinate
                    center_x = int(np.mean(cell_positions[1]))  # Calculate average x coordinate
                    cell_center = np.array([center_y, center_x])

                    # Calculate shortest distance to all separation points
                    distance, _ = tree.query(cell_center)

                    self.cells_info.at[index, 'leading_edge'] = distance

            else:
                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # Get cell number
                    cluster_label = cell_type_dict.get(cell_index, -1)  # Look up category from cell_type_dict

                    # If cell doesn't belong to current frame, skip
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue

                    # If category not in [0, 1], skip
                    if cluster_label not in [0, 1]:
                        continue

                    # Determine leading_edge sign based on cell type
                    if cluster_label == 1:  # Large cell
                        self.cells_info.at[index, 'leading_edge'] = -1
                    else:  # Small cell
                        self.cells_info.at[index, 'leading_edge'] = 1

            # Refresh image after drawing complete
            self.canvas.draw()

            # Save image using Pillow
            output_dir = self.output_picture
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            save_path = os.path.join(output_dir, f'{frame_number}_leading_edge.png')
            pil_img = Image.fromarray(img)
            pil_img.save(save_path)  # Save image using Pillow

    def update_image1(self):
        self.canvas.ax.clear()

        # Extract actual frame number from filename
        npy_file = self.npy_files[self.current_image_index]
        frame_number = int(re.findall(r'\d+', npy_file)[0])
        npy_path = os.path.join(self.npy_folder, npy_file)

        short_npy_name = os.path.basename(npy_path)
        self.log_info(f"Loading: {short_npy_name}", self.info_label)

        if os.path.exists(npy_path):
            # Load .npy file
            dat = np.load(npy_path, allow_pickle=True).item()

            # Extract outlines and colors data
            outlines = dat['outlines']  # Contour label map

            if 'img' in dat:
                base_img = dat['img']  # Use image from npy file
            else:
                # If 'img' not present, find corresponding image from img_folder matching frame_number
                img_file = None

                for file_name in os.listdir(self.img_folder):
                    # Extract numbers from image filename
                    img_frame_match = re.findall(r'\d+', file_name)
                    if img_frame_match:  # Check if numbers found
                        img_frame_number = int(img_frame_match[0])  # Extract first continuous number
                        # If image file frame matches npy file frame and is .png or .jpg
                        if img_frame_number == frame_number and (
                                file_name.endswith('.jpeg') or file_name.endswith('.jpg') or file_name.endswith('.png')):
                            img_file = os.path.join(self.img_folder, file_name)
                            break

                if img_file is not None:
                    # Open image and convert to numpy array
                    base_img = np.array(Image.open(img_file))
                else:
                    raise FileNotFoundError(f"Image file corresponding to frame {frame_number} not found!")

            # Convert grayscale image to RGB format
            if len(base_img.shape) == 2:
                base_img = np.stack([base_img] * 3, axis=-1)

            # Ensure image is uint8 format (avoid data clipping issues)
            if base_img.dtype != np.uint8:
                base_img = (base_img / base_img.max() * 255).astype(np.uint8)

            # Create a copy of RGB image
            img = base_img.copy()

            # Build mapping dictionary from Cell_Index to Cluster_Label
            cell_type_dict = self.clustering_data.set_index('Cell_Index')['Cluster_Label'].to_dict()

            # Store cell contours for adjacency analysis
            blue_cells_contours = {}  # Small cells (blue cells)
            yellow_cells_contours = {}  # Large cells (yellow cells)

            # Iterate through each unique contour ID and fill based on Excel category
            unique_ids = np.unique(outlines)  # Extract unique contour IDs
            for cell_id in unique_ids:
                if cell_id == 0:
                    continue  # Ignore background

                # Find category for current cell_id
                cell_index = f'Cell{frame_number}_{cell_id}'  # Build index in format Cell{frame_number}_{cell_id}
                cluster_label = cell_type_dict.get(cell_index, -1)  # Return -1 if category not found

                # Select fill color based on Cluster_Label and record cell contours
                if cluster_label == 1:
                    color = (255, 255, 0)  # Yellow (large cell)
                    yellow_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # Store large cell contour
                elif cluster_label == 0:
                    color = (0, 0, 255)  # Blue (small cell)
                    blue_cells_contours[cell_id] = np.array(np.where(outlines == cell_id)).T  # Store small cell contour
                else:
                    continue  # Skip other categories

                # Mark all pixel positions of this contour ID with specified color
                img[outlines == cell_id] = color

            # Find adjacent large and small cell pairs
            adjacent_pairs = self.get_adjacent_cells(blue_cells_contours, yellow_cells_contours, threshold=3)

            # Display image in Matplotlib
            ax = self.canvas.ax
            ax.imshow(img)
            ax.axis('off')  # Remove axes
            self.canvas.fig.subplots_adjust(left=0, right=1, top=1, bottom=0, wspace=0, hspace=0)
            ax.set_position([0, 0, 1, 1])

            # Collection of all separation points between adjacent large and small cells
            all_matching_points = []

            # Mark each adjacent cell pair and draw separation lines
            for large_cell_id, small_cell_id in adjacent_pairs:
                # Find adjacent matching points
                large_contour = yellow_cells_contours[large_cell_id]
                small_contour = blue_cells_contours[small_cell_id]

                # Get all matching points between this pair of adjacent cells
                matching_points = self.get_matching_points_between_cells(large_contour, small_contour, threshold=3)

                # Add all matching points to total point collection
                all_matching_points.extend([pt for large_point, pt in matching_points])

                # Draw separation lines between adjacent points
                for (large_point, small_point) in matching_points:
                    ax.plot([large_point[1], small_point[1]], [large_point[0], small_point[0]], color='red',
                            linestyle='--')

            # Add cell ID annotations
            for i, cell_id in enumerate(unique_ids):
                if cell_id == 0:
                    continue  # Ignore background

                # Find category for current cell_id, only label large and small cells
                cell_index = f'Cell{frame_number}_{cell_id}'
                cluster_label = cell_type_dict.get(cell_index, -1)

                # Only label large and small cells, skip other categories
                if cluster_label not in [0, 1]:
                    continue

                # Get all pixel positions of current cell (y, x)
                positions = np.where(outlines == cell_id)

                # Calculate cell center position
                center_y = int(np.mean(positions[0]))  # Calculate average y coordinate
                center_x = int(np.mean(positions[1]))  # Calculate average x coordinate

                # Set text box color based on cell category
                if cluster_label == 1:  # Large cell (yellow label)
                    bbox_props = dict(facecolor='yellow', alpha=0.5, edgecolor='none')
                elif cluster_label == 0:  # Small cell (blue label)
                    bbox_props = dict(facecolor='blue', alpha=0.5, edgecolor='none')
                else:
                    continue

                # Label cell ID at corresponding center position
                ax.text(center_x, center_y, str(cell_id), color='white', fontsize=7, bbox=bbox_props)

            # Use KDTree to calculate shortest distance from each cell in current frame to all matching points
            if all_matching_points:
                tree = KDTree(all_matching_points)

                # Iterate through all cells in cells_info, filtering for current frame cells
                for index, row in self.cells_info.iterrows():
                    cell_index = row['Cell Index']  # Get cell number
                    cluster_label = cell_type_dict.get(cell_index, -1)  # Look up category from cell_type_dict

                    # If cell doesn't belong to current frame, skip
                    if not cell_index.startswith(f'Cell{frame_number}_'):
                        continue

                    # If category not in [0, 1], skip
                    if cluster_label not in [0, 1]:
                        continue

                    # Calculate cell center point
                    cell_positions = np.where(outlines == int(cell_index.split('_')[-1]))
                    center_y = int(np.mean(cell_positions[0]))  # Calculate average y coordinate
                    center_x = int(np.mean(cell_positions[1]))  # Calculate average x coordinate
                    cell_center = np.array([center_y, center_x])

                    # Calculate shortest distance to all separation points
                    distance, _ = tree.query(cell_center)

                    self.update_cells_info(cell_index, distance)

            # Refresh image after drawing complete
            self.canvas.draw()

            # Save image using Pillow
            output_dir = self.output_picture
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)

            save_path = os.path.join(output_dir, f'{frame_number}_leading_edge.png')
            pil_img = Image.fromarray(img)
            pil_img.save(save_path)  # Save image using Pillow

    def on_click(self, event):
        if event.inaxes:
            for text in event.inaxes.texts:
                contains, _ = text.contains(event)
                if contains:
                    cell_id = text.get_text()
                    cell_index = f'Cell{self.current_image_index + 1}_{cell_id}'
                    dialog = CustomInputDialog(cell_index, self)
                    if dialog.exec_() == QDialog.Accepted:
                        new_cell_id, classification_result = dialog.getInputs()
                        # Update clustering data and save
                        self.update_clustering_data(cell_index, classification_result)
                        self.save_clustering_data()

                    break

    def update_clustering_data(self, cell_index, classification_result):
        row = self.clustering_data[self.clustering_data['Cell_Index'] == cell_index]
        if not row.empty:
            # Update classification result and mark row as modified
            if classification_result == 1:  # Update to large cell
                self.clustering_data.loc[
                    self.clustering_data['Cell_Index'] == cell_index, ['Cluster_Label', 'Cell_Type']] = [1, 'mes Cell']
            elif classification_result == 0:  # Update to small cell
                self.clustering_data.loc[
                    self.clustering_data['Cell_Index'] == cell_index, ['Cluster_Label', 'Cell_Type']] = [0, 'epi Cell']

            # Mark as modified
            self.clustering_data.loc[self.clustering_data['Cell_Index'] == cell_index, 'modified'] = True

    def save_clustering_data(self):
        try:
            self.update_image1()
            # Save modified rows
            self.save_modified_rows(self.clustering_data, self.clustering_data_path)

            self.save_modified_rows(self.cells_info, self.cells_info_path)
            QMessageBox.information(self, 'Save Successful', 'Changes have been saved to the file.')
        except Exception as e:
            QMessageBox.critical(self, 'Error', f'An error occurred while saving: {str(e)}')

    def update_cells_info(self, cell_index, new_value):
        # Check if 'leading_edge' column exists, add if not
        if 'leading_edge' not in self.cells_info.columns:
            self.cells_info['leading_edge'] = None  # Add new column with default value None

        # Find matching cell number
        row = self.cells_info[self.cells_info['Cell Index'] == cell_index]
        if not row.empty:
            # Update 'leading_edge' column value
            self.cells_info.loc[self.cells_info['Cell Index'] == cell_index, 'leading_edge'] = new_value

            # Add 'modified' column if not exists
            if 'modified' not in self.cells_info.columns:
                self.cells_info['modified'] = False  # Add new column with default value False

            # Mark as modified
            self.cells_info.loc[self.cells_info['Cell Index'] == cell_index, 'modified'] = True
